import pandas as pd
from openpyxl import load_workbook

file = "Data.xlsx"
wb = load_workbook(file)

constants = {
    '3D85C6': 'Very Good',
    'FFBED6EB': 'Good',
    'BFD7ED': 'Ok',
    'FFFFFFFF': 'Average',
    'FF000000': 'Average',
    'FFFF9999': 'Poor',
    'FF5555': 'Bad',
    'FF0000': 'Very Bad',
    '00000000': 'This shouldn"t be here',
    'FF9900': 'Solar',
    'FF9900FF': 'Legendary',
    'FF999999': 'Kinetic'
}

sheet_names = wb.sheetnames[3:len(wb.sheetnames)-1]

def get_cell_fill_color(cell):
    return cell.fill.start_color.index

limiter = 1

for sheet_name in sheet_names:
    sheet = wb[sheet_name]
    data = pd.read_excel(file, sheet_name=sheet_name, engine='openpyxl', header=1)
    data = data.iloc[0:49, 6:27]
    for row in sheet.iter_rows():
        for cell in row:
            print(get_cell_fill_color(cell))
            print(cell.value)
            print(constants[get_cell_fill_color(cell)])
    
    
    limiter += 1
    if limiter > 10:
        break